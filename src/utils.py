import mimetypes
import os
import signal
import shutil
import logging
import threading
import time
import json
from datetime import datetime, timedelta
from enum import Enum
from pathlib import Path
from typing import Annotated, List, Literal, Optional, Dict, Any, Union
from urllib.parse import quote

import boto3
from botocore.config import Config
import openpyxl
import sentry_sdk
from pydantic import BaseModel, Field, StringConstraints

logger = logging.getLogger(__name__)


def env_str(var: str, fallback: str) -> str:
    """`os.getenv` with empty string treated as unset.

    Compose passes `${VAR:-}` for optional fields, which arrives as "" not
    None — `os.getenv(var, fallback)` would then return "" instead of falling
    back. Using `or` collapses both unset and empty to the fallback.
    """
    return os.getenv(var) or fallback


def env_bool(var: str, fallback: bool) -> bool:
    """Parse a truthy/falsy env var. Empty/unset → fallback."""
    raw = os.getenv(var)
    if not raw:
        return fallback
    return raw.strip().lower() in {"1", "true", "yes", "y", "on"}


def env_int(var: str, fallback: int) -> int:
    """Parse an int env var. Empty/unset/unparseable → fallback."""
    raw = os.getenv(var)
    if not raw:
        return fallback
    try:
        return int(raw)
    except ValueError:
        return fallback


def get_calibrate_agent_cli() -> str:
    """Executable for the eval engine (PyPI package ``calibrate-agent``)."""
    return "calibrate-agent"


def capture_exception_to_sentry(exception: Exception) -> None:
    """
    Capture an exception to Sentry and mark it as unhandled.

    This ensures job failures appear as unresolved issues in Sentry
    rather than handled/resolved exceptions.
    """
    sentry_sdk.capture_exception(
        exception,
        hint={"mechanism": {"type": "generic", "handled": False}},
    )
    # Flush to ensure the event is sent immediately (important for background tasks)
    sentry_sdk.flush(timeout=2)


def build_tool_configs(agent_tools: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Build tool configurations for calibrate CLI from agent tools.

    Handles both structured output tools and webhook tools:
    - Structured output tools: Include type, name, description, and parameters
    - Webhook tools: Include type, name, description, and full webhook configuration

    Args:
        agent_tools: List of tool dicts from get_tools_for_agent()

    Returns:
        List of tool config dicts ready for calibrate config
    """
    tool_configs = []
    for tool in agent_tools:
        tool_config = tool.get("config", {})
        tool_type = tool_config.get("type", "structured_output")

        tool_entry = {
            "name": tool["name"],
            "description": tool["description"],
        }

        if tool_type == "webhook":
            # For webhook tools, include the full webhook configuration
            tool_entry["type"] = "webhook"
            tool_entry["parameters"] = tool_config.get("parameters", [])
            tool_entry["webhook"] = tool_config.get("webhook", {})
        else:
            # For structured output tools (default)
            tool_entry["type"] = "structured_output"
            tool_entry["parameters"] = tool_config.get("parameters", [])

        tool_configs.append(tool_entry)

    return tool_configs


# Timeout threshold for marking jobs as failed (5 minutes)
JOB_TIMEOUT_SECONDS = 3600
# Presigned URL caching constants
PRESIGNED_URL_EXPIRY_SECONDS = 3600  # 1 hour
PRESIGNED_URL_REFRESH_BUFFER_SECONDS = 300  # Refresh 5 minutes before expiry

# In-memory task storage (shared across routers)
tasks = {}
tasks_lock = threading.Lock()


# Lifecycle status for the run/job family (STT/TTS eval, agent tests,
# simulations, and annotation-eval jobs). `cancelled` is retained as a
# forward-compatible superset value; abort is tracked in `details.aborted`,
# not here. Kept as a `#` comment (not a docstring) so this internal note does
# NOT leak into the OpenAPI schema / public SDK — Pydantic promotes an enum's
# docstring to the schema `description`.
class TaskStatus(str, Enum):
    QUEUED = "queued"
    IN_PROGRESS = "in_progress"
    CANCELLED = "cancelled"
    DONE = "done"
    FAILED = "failed"


# Lifecycle status for the annotation family — labelling jobs
# (`pending → in_progress → completed`) and the public annotation-eval view
# (which normalizes the internal `done` to `completed`). `queued`/`failed` are
# included as a forward-compatible superset for eval jobs surfaced here.
# `#` comment (not a docstring) on purpose — see TaskStatus note above.
class AnnotationStatus(str, Enum):
    PENDING = "pending"
    QUEUED = "queued"
    IN_PROGRESS = "in_progress"
    COMPLETED = "completed"
    FAILED = "failed"


# Concrete value sets reused across routers for entity/job "type" fields.
EvaluatorTypeLiteral = Literal["tts", "stt", "llm", "llm-general", "conversation"]
DataTypeLiteral = Literal["text", "audio"]
OutputTypeLiteral = Literal["binary", "rating"]
EvaluatorKindLiteral = Literal["single", "side_by_side"]

# Shared field docs so evaluator_type / data_type / output_type read identically
# everywhere they appear (evaluators, public, simulations). Enum meanings render
# as a markdown bullet list.
EVALUATOR_TYPE_DESCRIPTION = (
    "What the evaluator judges:\n\n"
    "- `tts`: TTS audio\n"
    "- `stt`: one transcript\n"
    "- `llm`: a reply with its conversation history\n"
    "- `llm-general`: a standalone input and output pair\n"
    "- `conversation`: a full conversation\n"
)
DATA_TYPE_DESCRIPTION = (
    "The modality the judge reads:\n\n"
    "- `text`\n"
    "- `audio`\n"
)
OUTPUT_TYPE_DESCRIPTION = (
    "How the evaluator scores:\n\n"
    "- `binary`: pass or fail\n"
    "- `rating`: a numeric score, using the scale in `output_config`\n"
)
SimulationRunType = Literal["text", "voice"]
AgentTestJobType = Literal["llm-unit-test", "llm-benchmark"]
EvalJobType = Literal["stt-eval", "tts-eval", "annotation-eval"]
# Keep in sync with db.ANNOTATION_TASK_TYPES and db.VALID_EVALUATOR_TYPES
# (Literal requires literal members, so the vocabulary is mirrored here).
AnnotationTaskTypeLiteral = Literal["stt", "llm", "llm-general", "conversation"]
TestTypeLiteral = Literal["response", "tool_call", "conversation"]
MemberRoleLiteral = Literal["owner", "admin"]  # mirrors DB CHECK(role IN ('owner','admin'))
EvaluatorUuid = Annotated[str, StringConstraints(min_length=36, max_length=36)]

# Bulleted gloss of the agent `type` enum, shared across every model that
# exposes it (agents, agent-tools, agent-tests, simulations) so the two values
# read identically everywhere.
AGENT_TYPE_DESCRIPTION = (
    "- `agent`: built inside Calibrate\n"
    "- `connection`: your existing agent connected to Calibrate"
)
# Status a job can carry at *creation* time: it either starts immediately
# (`in_progress`) or waits for a concurrency slot (`queued`). Narrower than
# TaskStatus so create-response docs advertise only the reachable values.
InitialTaskStatus = Literal["queued", "in_progress"]

EXAMPLE_TEST_UUID = "b1c2d3e4-f5a6-7890-bcde-f12345678901"
# Bulleted gloss of the test `type` enum, shared by every model that exposes it
# (full test responses and the trimmed list shape) so the values read the same
# everywhere.
TEST_TYPE_DESCRIPTION = (
    "What the test judges:\n\n"
    "- `response`: judges the generated reply\n"
    "- `tool_call`: diffs the generated tool calls\n"
    "- `conversation`: judges the full conversation\n"
)


class TestListConfig(BaseModel):
    description: Optional[str] = Field(
        None,
        description="Short description of the test, shown in list views and searched on",
    )


class TestListResponse(BaseModel):
    """Trimmed test shape for list/index endpoints. The full config and hydrated
    evaluators live on the detail endpoint (`GET /tests/{uuid}`); the list keeps
    only what list and attach-dropdown views render."""

    uuid: str = Field(
        min_length=36,
        max_length=36,
        description="Unique ID for the test",
        examples=[EXAMPLE_TEST_UUID],
    )
    name: str = Field(description="Name of the test")
    type: TestTypeLiteral = Field(description=TEST_TYPE_DESCRIPTION)
    config: Optional[TestListConfig] = Field(
        None,
        description="Trimmed config carrying only the test's description. Fetch the test by ID for the full config and evaluators",
    )
    created_at: str = Field(
        description="When the test was created (ISO 8601 UTC)"
    )
    updated_at: str = Field(
        description="When the test was last updated (ISO 8601 UTC)"
    )


def to_test_list_response(test_dict: Dict[str, Any]) -> "TestListResponse":
    """Project a test row down to the trimmed list shape.

    Keeps only `config.description` and drops the heavy `config.history` /
    `evaluation` / `settings` blocks and evaluator hydration (an N+1). List and
    attach-dropdown views read only uuid/name/type/description; the edit and
    duplicate dialogs refetch the full test by ID."""
    config = test_dict.get("config")
    list_config = (
        TestListConfig(description=config.get("description"))
        if isinstance(config, dict)
        else None
    )
    return TestListResponse(
        uuid=test_dict["uuid"],
        name=test_dict["name"],
        type=test_dict["type"],
        config=list_config,
        created_at=test_dict["created_at"],
        updated_at=test_dict["updated_at"],
    )


class EvaluatorRunEntry(BaseModel):
    """One evaluator's aggregate metrics for an STT/TTS provider, paired with evaluator UUID.

    Canonical aggregate location: clients should read aggregate stats from
    here, not from the parallel ``provider_results[i].metrics[<display name>]``
    entries (those are kept for back-compat and may be retired in a later
    release).
    """

    evaluator_uuid: str = Field(
        min_length=36,
        max_length=36,
        description="ID of the evaluator",
    )
    metric_key: (
        str  # key as emitted in metrics.json (derived from CLI/config at run time)
    )
    aggregate: Dict[str, Any]
    name: Optional[str] = None  # filled on API read from DB + job snapshot
    description: Optional[str] = None  # filled on API read from current DB row
    evaluator_version_id: Optional[str] = Field(
        None,
        min_length=36,
        max_length=36,
        description="Pinned evaluator version ID at job-submit time",
    )
    output_type: Optional[OutputTypeLiteral] = Field(
        None, description="Output type. Drives per-row typing"
    )


class ProviderResult(BaseModel):
    provider: str
    success: Optional[bool] = None  # None while in progress, True/False when done
    metrics: Optional[Union[Dict[str, Any], List[Dict[str, Any]]]] = (
        None  # dict (new format) or list (backward compat)
    )
    results: Optional[List[Dict[str, Any]]] = None
    evaluator_runs: Optional[List[EvaluatorRunEntry]] = None


class TaskCreateResponse(BaseModel):
    task_id: str = Field(
        min_length=36,
        max_length=36,
        description="Unique identifier for this evaluation job",
        examples=["a3b2c1d0-e5f4-3210-abcd-ef1234567890"],
    )
    status: InitialTaskStatus = Field(
        description="Current status of the evaluation job"
    )
    dataset_id: Optional[str] = Field(
        None,
        min_length=36,
        max_length=36,
        description="ID of the dataset being evaluated",
        examples=["f47ac10b-58cc-4372-a567-0e02b2c3d479"],
    )
    dataset_name: Optional[str] = Field(
        None, description="Name of the dataset being evaluated"
    )


class TaskStatusResponse(BaseModel):
    task_id: str = Field(
        min_length=36,
        max_length=36,
        description="Evaluation job ID",
    )
    status: TaskStatus = Field(
        description="Current status of the evaluation job"
    )
    language: Optional[str] = None
    dataset_id: Optional[str] = Field(
        None,
        min_length=36,
        max_length=36,
        description="Source dataset ID",
    )
    dataset_name: Optional[str] = None
    provider_results: Optional[List[ProviderResult]] = None
    leaderboard_summary: Optional[List[Dict[str, Any]]] = None
    error: Optional[str] = None
    is_public: bool = False
    share_token: Optional[str] = None


def kill_process_group(pid: int, job_id: str) -> bool:
    """Kill a process group by PID.

    Args:
        pid: Process ID (also used as PGID when start_new_session=True)
        job_id: Job ID for logging

    Returns:
        True if process was killed or didn't exist, False on error
    """
    if not pid:
        return True

    try:
        os.killpg(pid, signal.SIGTERM)
        logger.info(f"Job {job_id}: Sent SIGTERM to process group {pid}")

        time.sleep(0.5)

        try:
            os.killpg(pid, signal.SIGKILL)
            logger.info(f"Job {job_id}: Sent SIGKILL to process group {pid}")
        except (ProcessLookupError, PermissionError):
            # Process already terminated after SIGTERM
            logger.info(
                f"Job {job_id}: Process group {pid} already terminated after SIGTERM"
            )

        return True
    except ProcessLookupError:
        logger.info(f"Job {job_id}: Process group {pid} not found (already dead)")
        return True
    except PermissionError:
        logger.warning(f"Job {job_id}: No permission to kill process group {pid}")
        return False
    except Exception as e:
        logger.error(f"Job {job_id}: Error killing process group {pid}: {e}")
        return False


def kill_processes_from_dict(pids_dict: dict, job_id: str) -> None:
    """Kill multiple processes from a dict mapping (e.g., provider -> PID).

    Args:
        pids_dict: Dict mapping names to PIDs (e.g., {"deepgram": 12345, "openai": 12346})
        job_id: Job ID for logging
    """
    if not pids_dict:
        logger.info(f"Job {job_id}: No running PIDs to kill")
        return

    for name, pid in pids_dict.items():
        if not pid:
            continue
        try:
            os.killpg(pid, signal.SIGTERM)
            logger.info(f"Job {job_id}: Sent SIGTERM to process group {pid} ({name})")

            time.sleep(0.5)

            try:
                os.killpg(pid, signal.SIGKILL)
                logger.info(
                    f"Job {job_id}: Sent SIGKILL to process group {pid} ({name})"
                )
            except ProcessLookupError:
                logger.info(
                    f"Job {job_id}: Process group {pid} ({name}) already terminated"
                )
        except ProcessLookupError:
            logger.info(f"Job {job_id}: Process group {pid} ({name}) not found")
        except PermissionError:
            logger.warning(
                f"Job {job_id}: No permission to kill process group {pid} ({name})"
            )
        except Exception as e:
            logger.error(
                f"Job {job_id}: Error killing process group {pid} ({name}): {e}"
            )


def is_job_timed_out(
    updated_at: str, timeout_seconds: int = JOB_TIMEOUT_SECONDS
) -> bool:
    """Check if a job has timed out based on its updated_at timestamp.

    Args:
        updated_at: ISO format timestamp string (from SQLite, stored in UTC)

    Returns:
        True if the job hasn't been updated in more than JOB_TIMEOUT_SECONDS
    """
    try:
        # Parse the timestamp (SQLite format: "YYYY-MM-DD HH:MM:SS", stored in UTC)
        last_update = datetime.fromisoformat(updated_at.replace(" ", "T"))
        # Use UTC for comparison since SQLite CURRENT_TIMESTAMP is in UTC
        timeout_threshold = datetime.utcnow() - timedelta(seconds=timeout_seconds)
        return last_update < timeout_threshold
    except Exception as e:
        logger.warning(f"Error parsing timestamp {updated_at}: {e}")
        return False


def get_s3_client():
    """Get S3-compatible client. Honors S3_ENDPOINT_URL for GCS interop.

    Treats empty strings as unset so docker-compose passing through
    `${AWS_REGION:-}` (etc.) doesn't override the code default with "".

    When S3_ENDPOINT_URL is set (i.e. talking to a non-AWS endpoint like GCS),
    pin checksum behavior to "when_required" — boto3 >=1.36's default
    "when_supported" adds x-amz-checksum-* headers that GCS's S3 interop layer
    rejects with SignatureDoesNotMatch. AWS S3 itself handles either setting.
    """
    # Reuse the validating mode check so an invalid OBJECT_STORAGE_MODE fails the
    # same way here as everywhere else (rather than silently building a real
    # client for a typo'd value).
    if is_local_object_storage():
        return None

    endpoint_url = os.getenv("S3_ENDPOINT_URL") or None
    aws_access_key_id = os.getenv("AWS_ACCESS_KEY_ID") or None
    aws_secret_access_key = os.getenv("AWS_SECRET_ACCESS_KEY") or None
    aws_region = os.getenv("AWS_REGION") or "ap-south-1"

    kwargs = {"region_name": aws_region}
    if endpoint_url:
        kwargs["endpoint_url"] = endpoint_url
        kwargs["config"] = Config(
            request_checksum_calculation="when_required",
            response_checksum_validation="when_required",
        )
    if aws_access_key_id and aws_secret_access_key:
        kwargs["aws_access_key_id"] = aws_access_key_id
        kwargs["aws_secret_access_key"] = aws_secret_access_key

    return boto3.client("s3", **kwargs)


LOCAL_STORAGE_BUCKET = "local-dev-artifacts"


def get_object_storage_mode() -> str:
    """Return the configured artifact storage mode.

    ``s3`` is the production default. ``local`` is a development-only shim that
    stores objects under ``LOCAL_ARTIFACT_ROOT`` using the same object keys.
    """
    mode = (os.getenv("OBJECT_STORAGE_MODE") or "s3").strip().lower()
    if mode not in {"s3", "local"}:
        raise ValueError("OBJECT_STORAGE_MODE must be either 's3' or 'local'")
    return mode


def is_local_object_storage() -> bool:
    return get_object_storage_mode() == "local"


def get_local_artifact_root() -> Path:
    configured = os.getenv("LOCAL_ARTIFACT_ROOT")
    if configured:
        root = Path(configured)
    else:
        root = Path(env_str("DB_ROOT_DIR", ".")) / "artifacts"
    return root.expanduser().resolve()


def get_local_artifact_path(key: str) -> Path:
    root = get_local_artifact_root()
    parts = [part for part in key.replace("\\", "/").split("/") if part]
    if not parts or any(part in {".", ".."} for part in parts):
        raise ValueError("Invalid local artifact path")
    path = (root.joinpath(*parts)).resolve()
    path.relative_to(root)
    return path


def get_local_artifact_url(key: str) -> str:
    path = quote(key.lstrip("/"), safe="/")
    relative_url = f"/local-artifacts/{path}"
    base_url = os.getenv("LOCAL_ARTIFACT_BASE_URL")
    if not base_url:
        return relative_url
    return f"{base_url.rstrip('/')}{relative_url}"


def get_s3_output_config():
    """Get the configured artifact bucket.

    In local development mode this returns a stable sentinel bucket so existing
    ``s3://bucket/key`` paths keep working without AWS.
    """
    if is_local_object_storage():
        return os.getenv("LOCAL_ARTIFACT_BUCKET") or LOCAL_STORAGE_BUCKET

    bucket = os.getenv("S3_OUTPUT_BUCKET")

    if not bucket:
        raise ValueError("S3_OUTPUT_BUCKET environment variable is required")

    return bucket


def upload_file_to_s3(
    s3_client,
    local_path: Union[str, Path],
    bucket: str,
    s3_key: str,
) -> None:
    """Upload a local file to S3 with a guessed Content-Type.

    Setting Content-Type ensures browsers render supported files (JSON, audio,
    text) inline instead of forcing a download (which happens when S3 defaults
    to binary/octet-stream).
    """
    if is_local_object_storage():
        destination = get_local_artifact_path(s3_key)
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(local_path, destination)
        logger.info("Stored local artifact at %s", destination)
        return

    content_type, _ = mimetypes.guess_type(str(local_path))
    extra_args = {"ContentType": content_type} if content_type else None
    s3_client.upload_file(str(local_path), bucket, s3_key, ExtraArgs=extra_args)


def download_file_from_s3(
    s3_client,
    bucket: str,
    s3_key: str,
    local_path: Union[str, Path],
) -> None:
    """Download a single object to ``local_path``.

    In local mode this copies from ``LOCAL_ARTIFACT_ROOT`` using the same key
    (``bucket`` is ignored, mirroring :func:`upload_file_to_s3`), so job runners
    can fetch their inputs without a real S3 client.
    """
    if is_local_object_storage():
        source = get_local_artifact_path(s3_key)
        Path(local_path).parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, local_path)
        return

    s3_client.download_file(bucket, s3_key, str(local_path))


def upload_top_level_files_to_s3(
    s3_client,
    local_dir: Path,
    bucket: str,
    key_prefix: str,
) -> None:
    """Upload only regular files directly under ``local_dir`` (e.g. run-level ``logs``, ``leaderboard.csv``).

    Subdirectories are ignored. Use :func:`upload_directory_tree_to_s3` for a full tree.
    """
    if not local_dir or not local_dir.is_dir():
        return
    prefix = key_prefix.rstrip("/")
    for p in local_dir.iterdir():
        if p.is_file():
            s3_key = f"{prefix}/{p.name}"
            upload_file_to_s3(s3_client, p, bucket, s3_key)


def upload_directory_tree_to_s3(
    s3_client,
    local_root: Path,
    bucket: str,
    key_prefix: str,
) -> None:
    """Recursively upload every file under ``local_root`` to ``s3://bucket/{key_prefix}/<relative>``."""
    if not local_root or not local_root.exists():
        return
    local_root = local_root.resolve()
    prefix = key_prefix.rstrip("/")
    for root, dirs, files in os.walk(local_root):
        for file in files:
            local_file_path = Path(root) / file
            relative_path = local_file_path.relative_to(local_root)
            s3_key = f"{prefix}/{relative_path.as_posix()}"
            upload_file_to_s3(s3_client, local_file_path, bucket, s3_key)


def list_object_keys(s3_client, bucket: str, prefix: str = "") -> List[str]:
    """List object keys under ``prefix``.

    In local mode this walks ``LOCAL_ARTIFACT_ROOT`` and returns keys relative to
    it (``bucket`` ignored, mirroring the upload/download helpers), so listing
    works without a real S3 client. In S3 mode it paginates ``list_objects_v2``.
    """
    if is_local_object_storage():
        root = get_local_artifact_root()
        base = get_local_artifact_path(prefix) if prefix.strip("/") else root
        if not base.exists():
            return []
        return [
            p.relative_to(root).as_posix()
            for p in sorted(base.rglob("*"))
            if p.is_file()
        ]

    keys: List[str] = []
    paginator = s3_client.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        for obj in page.get("Contents", []):
            keys.append(obj["Key"])
    return keys


def generate_presigned_download_url(
    s3_key: str,
    bucket: Optional[str] = None,
    expiration: int = PRESIGNED_URL_EXPIRY_SECONDS,
) -> Optional[str]:
    """Generate a presigned URL for downloading (get_object) from S3.

    Args:
        s3_key: The S3 object key
        bucket: S3 bucket name (defaults to S3_OUTPUT_BUCKET env var)
        expiration: URL expiration time in seconds (default: 1 hour)

    Returns:
        Presigned URL string, or None if generation fails
    """
    try:
        if is_local_object_storage():
            return get_local_artifact_url(s3_key)

        s3 = get_s3_client()
        s3_bucket = bucket or get_s3_output_config()

        return s3.generate_presigned_url(
            "get_object",
            Params={
                "Bucket": s3_bucket,
                "Key": s3_key,
            },
            ExpiresIn=expiration,
        )
    except Exception as e:
        logger.warning(f"Failed to generate presigned download URL for {s3_key}: {e}")
        return None


def presign_audio_path(audio_path: Optional[str]) -> Optional[str]:
    """Convert an s3://bucket/key path (or plain key) to a presigned download URL."""
    if not audio_path:
        return audio_path
    if audio_path.startswith("s3://"):
        parts = audio_path[5:].split("/", 1)
        bucket = parts[0]
        key = parts[1] if len(parts) > 1 else ""
        return generate_presigned_download_url(key, bucket=bucket) or audio_path
    if audio_path.startswith("http"):
        return audio_path
    return generate_presigned_download_url(audio_path) or audio_path


def generate_presigned_upload_url(
    s3_key: str,
    content_type: str,
    bucket: Optional[str] = None,
    expiration: int = PRESIGNED_URL_EXPIRY_SECONDS,
) -> Optional[str]:
    """Generate a presigned URL for uploading (put_object) to S3.

    Args:
        s3_key: The S3 object key
        content_type: The content type of the file to upload
        bucket: S3 bucket name (defaults to S3_OUTPUT_BUCKET env var)
        expiration: URL expiration time in seconds (default: 1 hour)

    Returns:
        Presigned URL string, or None if generation fails
    """
    try:
        if is_local_object_storage():
            return get_local_artifact_url(s3_key)

        s3 = get_s3_client()
        s3_bucket = bucket or get_s3_output_config()

        return s3.generate_presigned_url(
            "put_object",
            Params={
                "Bucket": s3_bucket,
                "Key": s3_key,
                "ContentType": content_type,
            },
            ExpiresIn=expiration,
        )
    except Exception as e:
        logger.warning(f"Failed to generate presigned upload URL for {s3_key}: {e}")
        return None


def get_max_concurrent_jobs() -> int:
    """Get the maximum number of concurrent jobs from environment variable.

    Defaults to 2 if not set.
    """
    return int(os.getenv("MAX_CONCURRENT_JOBS"))


def get_max_concurrent_jobs_per_org() -> int:
    """Get the maximum number of concurrent jobs per org from environment variable.

    Defaults to 1 if not set. Set to 0 to disable org-level limit.

    For one release we read `MAX_CONCURRENT_JOBS_PER_ORG` (new name) and fall back
    to `MAX_CONCURRENT_JOBS_PER_USER` (old name) — the migration renamed the
    scope from per-user to per-org. The old env var keeps working so deployments
    can roll forward without a coordinated config change.
    """
    raw_new = os.getenv("MAX_CONCURRENT_JOBS_PER_ORG")
    if raw_new is not None:
        return int(raw_new)
    raw_old = os.getenv("MAX_CONCURRENT_JOBS_PER_USER")
    if raw_old is not None:
        logger.warning(
            "MAX_CONCURRENT_JOBS_PER_USER is deprecated; rename to "
            "MAX_CONCURRENT_JOBS_PER_ORG."
        )
        return int(raw_old)
    return 1


# Job queue lock to ensure thread-safe queue operations
_job_queue_lock = threading.Lock()

# Registry of job starter callbacks by job type
_job_starters: Dict[str, callable] = {}


def register_job_starter(job_type: str, starter_callback: callable) -> None:
    """Register a callback function for starting jobs of a specific type.

    Args:
        job_type: The job type (e.g., "stt-eval", "tts-eval")
        starter_callback: Function that takes a job dict and starts the job.
    """
    _job_starters[job_type] = starter_callback
    logger.info(f"Registered job starter for type: {job_type}")


def try_start_queued_job(job_types: List[str]) -> bool:
    """Try to start the next queued job if there's capacity.

    Checks both global limit and per-org limit for each queued job.

    Args:
        job_types: List of job types to consider (e.g., ["stt-eval", "tts-eval"])

    Returns:
        True if a job was started, False otherwise.
    """
    from db import (
        count_running_jobs,
        count_running_jobs_for_org,
        get_queued_jobs,
        update_job,
    )

    with _job_queue_lock:
        max_jobs = get_max_concurrent_jobs()
        running_count = count_running_jobs(job_types)

        logger.info(f"Job queue check: {running_count}/{max_jobs} jobs running")

        if running_count >= max_jobs:
            logger.info("Max concurrent jobs reached, not starting new job")
            return False

        queued_jobs = get_queued_jobs(job_types)
        if not queued_jobs:
            logger.info("No queued jobs to start")
            return False

        max_jobs_per_org = get_max_concurrent_jobs_per_org()
        job_to_start = None

        for job in queued_jobs:
            org_uuid = job.get("org_uuid")
            if max_jobs_per_org > 0 and org_uuid:  # 0 means disabled
                org_running_count = count_running_jobs_for_org(org_uuid, job_types)
                if org_running_count >= max_jobs_per_org:
                    logger.info(
                        f"Org {org_uuid} has {org_running_count}/{max_jobs_per_org} jobs running, skipping job {job['uuid']}"
                    )
                    continue
            job_to_start = job
            break

        if not job_to_start:
            logger.info("No queued jobs can be started (all orgs at their limit)")
            return False

        job_id = job_to_start["uuid"]
        job_type = job_to_start.get("type")

        # Find the appropriate starter callback
        starter_callback = _job_starters.get(job_type)
        if not starter_callback:
            logger.error(f"No job starter registered for type: {job_type}")
            return False

        # Update status to in_progress before starting
        update_job(job_id, status=TaskStatus.IN_PROGRESS.value)
        logger.info(f"Starting queued job {job_id} of type {job_type}")

        try:
            # Start the job (this should spawn a thread)
            starter_callback(job_to_start)
            return True
        except Exception as e:
            # If starting fails, mark as done with error
            logger.error(f"Failed to start job {job_id}: {e}")
            update_job(
                job_id,
                status=TaskStatus.DONE.value,
                results={"error": f"Failed to start job: {str(e)}"},
            )
            return False


def can_start_job(job_types: List[str], org_uuid: str) -> bool:
    """Check if there's capacity to start a new job immediately.

    Checks both global limit and per-org limit.

    Args:
        job_types: List of job types to consider for counting running jobs.
        org_uuid: UUID of the org requesting the job.

    Returns:
        True if a new job can be started, False otherwise.
    """
    from db import count_running_jobs, count_running_jobs_for_org

    with _job_queue_lock:
        max_jobs = get_max_concurrent_jobs()
        running_count = count_running_jobs(job_types)
        if running_count >= max_jobs:
            return False

        max_jobs_per_org = get_max_concurrent_jobs_per_org()
        if max_jobs_per_org > 0:  # 0 means disabled
            org_running_count = count_running_jobs_for_org(org_uuid, job_types)
            if org_running_count >= max_jobs_per_org:
                return False

        return True


# ============ Agent Test Job Queue Functions ============


def try_start_queued_agent_test_job(job_types: List[str]) -> bool:
    """Try to start the next queued agent test job if there's capacity.

    Checks both global limit and per-org limit for each queued job.

    Args:
        job_types: List of job types to consider (e.g., ["llm-unit-test", "llm-benchmark"])

    Returns:
        True if a job was started, False otherwise.
    """
    from db import (
        count_running_agent_test_jobs,
        count_running_agent_test_jobs_for_org,
        get_queued_agent_test_jobs,
        update_agent_test_job,
    )

    with _job_queue_lock:
        max_jobs = get_max_concurrent_jobs()
        running_count = count_running_agent_test_jobs(job_types)

        logger.info(
            f"Agent test job queue check: {running_count}/{max_jobs} jobs running"
        )

        if running_count >= max_jobs:
            logger.info("Max concurrent jobs reached, not starting new agent test job")
            return False

        queued_jobs = get_queued_agent_test_jobs(job_types)
        if not queued_jobs:
            logger.info("No queued agent test jobs to start")
            return False

        max_jobs_per_org = get_max_concurrent_jobs_per_org()
        job_to_start = None

        for job in queued_jobs:
            org_uuid = job.get("org_uuid")
            if max_jobs_per_org > 0 and org_uuid:  # 0 means disabled
                org_running_count = count_running_agent_test_jobs_for_org(
                    org_uuid, job_types
                )
                if org_running_count >= max_jobs_per_org:
                    logger.info(
                        f"Org {org_uuid} has {org_running_count}/{max_jobs_per_org} agent test jobs running, skipping job {job['uuid']}"
                    )
                    continue
            job_to_start = job
            break

        if not job_to_start:
            logger.info(
                "No queued agent test jobs can be started (all orgs at their limit)"
            )
            return False

        job_id = job_to_start["uuid"]
        job_type = job_to_start.get("type")

        # Find the appropriate starter callback
        starter_callback = _job_starters.get(job_type)
        if not starter_callback:
            logger.error(f"No job starter registered for type: {job_type}")
            return False

        # Update status to in_progress before starting
        update_agent_test_job(job_id, status=TaskStatus.IN_PROGRESS.value)
        logger.info(f"Starting queued agent test job {job_id} of type {job_type}")

        try:
            # Start the job (this should spawn a thread)
            starter_callback(job_to_start)
            return True
        except Exception as e:
            # If starting fails, mark as done with error
            logger.error(f"Failed to start agent test job {job_id}: {e}")
            update_agent_test_job(
                job_id,
                status=TaskStatus.DONE.value,
                results={"error": f"Failed to start job: {str(e)}"},
            )
            return False


def can_start_agent_test_job(job_types: List[str], org_uuid: str) -> bool:
    """Check if there's capacity to start a new agent test job immediately.

    Checks both global limit and per-org limit.

    Args:
        job_types: List of job types to consider for counting running jobs.
        org_uuid: UUID of the org requesting the job.

    Returns:
        True if a new job can be started, False otherwise.
    """
    from db import count_running_agent_test_jobs, count_running_agent_test_jobs_for_org

    with _job_queue_lock:
        max_jobs = get_max_concurrent_jobs()
        running_count = count_running_agent_test_jobs(job_types)
        if running_count >= max_jobs:
            return False

        max_jobs_per_org = get_max_concurrent_jobs_per_org()
        if max_jobs_per_org > 0:  # 0 means disabled
            org_running_count = count_running_agent_test_jobs_for_org(
                org_uuid, job_types
            )
            if org_running_count >= max_jobs_per_org:
                return False

        return True


# ============ Simulation Job Queue Functions ============


def try_start_queued_simulation_job(job_types: List[str]) -> bool:
    """Try to start the next queued simulation job if there's capacity.

    Checks both global limit and per-org limit for each queued job.

    Args:
        job_types: List of job types to consider (e.g., ["text", "voice"])

    Returns:
        True if a job was started, False otherwise.
    """
    from db import (
        count_running_simulation_jobs,
        count_running_simulation_jobs_for_org,
        get_queued_simulation_jobs,
        update_simulation_job,
    )

    with _job_queue_lock:
        max_jobs = get_max_concurrent_jobs()
        running_count = count_running_simulation_jobs(job_types)

        logger.info(
            f"Simulation job queue check: {running_count}/{max_jobs} jobs running"
        )

        if running_count >= max_jobs:
            logger.info("Max concurrent jobs reached, not starting new simulation job")
            return False

        queued_jobs = get_queued_simulation_jobs(job_types)
        if not queued_jobs:
            logger.info("No queued simulation jobs to start")
            return False

        max_jobs_per_org = get_max_concurrent_jobs_per_org()
        job_to_start = None

        for job in queued_jobs:
            org_uuid = job.get("org_uuid")
            if max_jobs_per_org > 0 and org_uuid:  # 0 means disabled
                org_running_count = count_running_simulation_jobs_for_org(
                    org_uuid, job_types
                )
                if org_running_count >= max_jobs_per_org:
                    logger.info(
                        f"Org {org_uuid} has {org_running_count}/{max_jobs_per_org} simulation jobs running, skipping job {job['uuid']}"
                    )
                    continue
            job_to_start = job
            break

        if not job_to_start:
            logger.info(
                "No queued simulation jobs can be started (all orgs at their limit)"
            )
            return False

        job_id = job_to_start["uuid"]
        job_type = job_to_start.get("type")

        # Find the appropriate starter callback
        starter_callback = _job_starters.get(job_type)
        if not starter_callback:
            logger.error(f"No job starter registered for type: {job_type}")
            return False

        # Update status to in_progress before starting
        update_simulation_job(job_id, status=TaskStatus.IN_PROGRESS.value)
        logger.info(f"Starting queued simulation job {job_id} of type {job_type}")

        try:
            # Start the job (this should spawn a thread)
            starter_callback(job_to_start)
            return True
        except Exception as e:
            # If starting fails, mark as done with error
            logger.error(f"Failed to start simulation job {job_id}: {e}")
            update_simulation_job(
                job_id,
                status=TaskStatus.DONE.value,
                results={"error": f"Failed to start job: {str(e)}"},
            )
            return False


def can_start_simulation_job(job_types: List[str], org_uuid: str) -> bool:
    """Check if there's capacity to start a new simulation job immediately.

    Checks both global limit and per-org limit.

    Args:
        job_types: List of job types to consider for counting running jobs.
        org_uuid: UUID of the org requesting the job.

    Returns:
        True if a new job can be started, False otherwise.
    """
    from db import count_running_simulation_jobs, count_running_simulation_jobs_for_org

    with _job_queue_lock:
        max_jobs = get_max_concurrent_jobs()
        running_count = count_running_simulation_jobs(job_types)
        if running_count >= max_jobs:
            return False

        max_jobs_per_org = get_max_concurrent_jobs_per_org()
        if max_jobs_per_org > 0:  # 0 means disabled
            org_running_count = count_running_simulation_jobs_for_org(
                org_uuid, job_types
            )
            if org_running_count >= max_jobs_per_org:
                return False

        return True


def normalize_metrics(metrics):
    """Convert old list-of-dicts metrics format to new dict format.

    Old format: [{"wer": 2.4}, {"string_similarity": 0.15}, {"metric_name": "ttfb", "mean": 0.1, ...}, ...]
    New format: {"wer": 2.4, "string_similarity": 0.15, "ttfb": {"mean": 0.1, ...}, ...}
    """
    if metrics is None:
        return None
    if isinstance(metrics, dict):
        return metrics
    if isinstance(metrics, list):
        result = {}
        for item in metrics:
            if isinstance(item, dict):
                if "metric_name" in item:
                    metric_name = item["metric_name"]
                    value = {k: v for k, v in item.items() if k != "metric_name"}
                    result[metric_name] = value
                else:
                    result.update(item)
        return result if result else metrics
    return metrics


def is_evaluator_metric_aggregate(value: Any) -> bool:
    """True for nested evaluator outputs in metrics.json (excludes wer scalars, ttfb, etc.)."""
    return isinstance(value, dict) and "type" in value


def ordered_evaluator_metric_keys(metrics: Optional[Dict[str, Any]]) -> List[str]:
    if not metrics or not isinstance(metrics, dict):
        return []
    return [k for k, v in metrics.items() if is_evaluator_metric_aggregate(v)]


def read_evaluators_map_from_config(output_dir: Optional[Path]) -> Dict[str, str]:
    """Read root config.json evaluators_map as {metric_key: evaluator_uuid}."""
    if not output_dir:
        return {}
    config_path = output_dir / "config.json"
    if not config_path.exists():
        return {}
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
    except Exception as e:
        logger.warning(f"Failed to read evaluator map from {config_path}: {e}")
        return {}

    raw = config.get("evaluators_map")
    if not isinstance(raw, dict):
        return {}
    return {
        str(name): str(evaluator_id)
        for evaluator_id, name in raw.items()
        if evaluator_id and name
    }


def build_evaluator_runs_for_eval_job(
    metrics: Any,
    evaluator_id_by_metric_key: Optional[Dict[str, str]] = None,
) -> List[EvaluatorRunEntry]:
    """Pair evaluator aggregates with UUIDs from calibrate's config map."""
    metrics_dict = normalize_metrics(metrics)
    if not isinstance(metrics_dict, dict):
        return []
    keys = ordered_evaluator_metric_keys(metrics_dict)
    if not keys:
        return []
    evaluator_id_by_metric_key = evaluator_id_by_metric_key or {}
    runs: List[EvaluatorRunEntry] = []
    for mk in keys:
        eu = evaluator_id_by_metric_key.get(mk)
        if not eu:
            continue
        agg = metrics_dict.get(mk)
        if agg is None:
            continue
        runs.append(
            EvaluatorRunEntry(
                evaluator_uuid=eu,
                metric_key=mk,
                aggregate=dict(agg) if isinstance(agg, dict) else {"value": agg},
            )
        )
    return runs


# Coerced as floats. Non-evaluator columns calibrate writes as strings into
# results.csv but that are numeric by contract. Conservative list — anything
# unknown is left as a string so we don't silently drop a future text column.
_NUMERIC_ROW_KEYS = frozenset(
    {
        "wer",
        "cer",
        "string_similarity",
        "similarity",
        "processing_time",
        "ttfb",
        "latency",
        "duration",
        "audio_duration",
    }
)


def _coerce_numeric(value: Any) -> Any:
    """Return float(value) when the string parses as numeric. Otherwise unchanged."""
    if value is None or isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, str):
        s = value.strip()
        if not s or s.lower() in {"nan", "none", "null"}:
            return None
        try:
            f = float(s)
            return f
        except ValueError:
            return value
    return value


def coerce_evaluator_score(raw: Any, output_type: str) -> Any:
    """Coerce a raw evaluator value out of CSV/JSON into the right Python
    type per ``output_type``. Falls back to passthrough on unparseable input.

    Single source of truth shared by the API-facing post-processor
    (`post_process_provider_results`) and the annotation-eval persistence
    path (`annotation_eval_runner._row_evaluator_value`). Previously these
    had divergent implementations — annotation-eval handled stringified
    floats like ``"1.0"`` / ``"0.0"`` (which calibrate's simulation flow
    emits), the API path didn't, so binary evaluators rendered as the
    raw string instead of a bool on the FE. Unified here.

    Binary handling has to cope with the full range of representations
    calibrate emits across its three flows: bool, ``"True"``/``"False"``
    strings, ``"1"``/``"0"`` strings, ``"1.0"``/``"0.0"`` stringified
    floats, and bare numerics.

    Rating returns ``int`` when the value is whole-numbered (the common
    case — 1-5 / 1-10 scales), ``float`` when fractional, preserving
    precision for any future fractional rating scale.
    """
    if output_type == "binary":
        if isinstance(raw, bool):
            return raw
        if isinstance(raw, (int, float)):
            return bool(raw)
        s = str(raw).strip().lower()
        if s in ("true", "yes", "pass", "passed"):
            return True
        if s in ("false", "no", "fail", "failed"):
            return False
        # Numeric strings — covers "1", "0", "1.0", "0.0", "1.00", etc.
        try:
            return bool(float(s))
        except (TypeError, ValueError):
            return raw
    if output_type == "rating":
        try:
            f = float(raw)
        except (TypeError, ValueError):
            return raw
        return int(f) if f.is_integer() else f
    return raw


def _row_value_looks_like_error(raw_value: Any, reasoning: Any) -> bool:
    """Heuristic: did calibrate write an error marker instead of a real judgement?
    Patterns observed: literal string "ERROR" in the value cell, or empty value
    with reasoning starting with 'Error' / 'Exception'."""
    if isinstance(raw_value, str) and raw_value.strip().lower() in {"error", "err"}:
        return True
    if (raw_value is None or raw_value == "") and isinstance(reasoning, str):
        head = reasoning.strip().lower()[:20]
        if head.startswith("error") or head.startswith("exception"):
            return True
    return False


def post_process_provider_results(
    provider_results: Optional[List[Dict[str, Any]]],
    evaluator_snapshots: Optional[List[Dict[str, Any]]] = None,
    evaluator_id_by_metric_key: Optional[Dict[str, str]] = None,
) -> None:
    """Single canonical post-processor for STT/TTS provider_results before they
    leave the API. Idempotent — safe to call on already-shaped data.

    Mutations performed in place:

    1. ``evaluator_runs`` populated per-provider as soon as that provider's
       ``metrics.json`` is on disk (the in-progress path historically only
       set this once the *whole* run finished). Built from
       ``ordered_evaluator_metric_keys`` + the metric_key→uuid map.
    2. Each ``evaluator_runs`` entry gets ``evaluator_version_id`` and
       ``output_type`` filled from the job-time evaluator snapshot, so the
       FE can render rubric-aware widgets without a second DB roundtrip.
    3. Per-row outputs are namespaced under ``row["evaluator_outputs"][uuid]``
       (``{value, reasoning, version_id, output_type, error?}``) by lifting
       the flat ``<column_name>`` / ``<column_name>_reasoning`` keys.
       Legacy flat keys are KEPT for one deprecation window — clients that
       have migrated read from ``evaluator_outputs`` only.

       The lift is driven by ``evaluator_id_by_metric_key`` (calibrate's
       authoritative ``config.json/evaluators_map``) when present — that's
       the only string we *know* identifies an evaluator output column.
       Falls back to the snapshot's display name when the map is missing
       (in-progress jobs whose config.json hasn't landed, or pre-migration
       legacy data). The fallback path is vulnerable to collisions with
       built-in row columns (``id``, ``gt``, ``pred``, ``wer``, ...) and
       with duplicate evaluator names, so it's a back-compat path only.
    4. Values are typed: binary → ``bool``, rating → numeric. Known numeric
       row columns (wer/cer/string_similarity/processing_time/etc.) are
       parsed from string to ``float`` here too.
    5. Per-row judge failures (calibrate wrote ``"ERROR"`` instead of a real
       value, or left it blank with an ``"Error: ..."`` reasoning) surface
       as ``evaluator_outputs[uuid].error = True`` with the original message
       preserved in ``.reasoning``.
    """
    if not provider_results:
        return

    snapshot_by_uuid: Dict[str, Dict[str, Any]] = {}
    snapshot_by_name: Dict[str, Dict[str, Any]] = {}
    for snap in evaluator_snapshots or []:
        if not isinstance(snap, dict):
            continue
        uid = snap.get("uuid") or snap.get("evaluator_id")
        if uid:
            snapshot_by_uuid[str(uid)] = snap
        nm = snap.get("name")
        if nm:
            # Calibrate writes columns under the *rendered* display name. Index
            # by name so we can recover the UUID when lifting flat row keys.
            snapshot_by_name[str(nm)] = snap

    evaluator_id_by_metric_key = evaluator_id_by_metric_key or {}

    for pr in provider_results:
        # ---------- Step 1 + 5: evaluator_runs from metrics.json ----------
        runs = pr.get("evaluator_runs")
        metrics = pr.get("metrics")
        if not runs and metrics is not None and evaluator_id_by_metric_key:
            try:
                built = build_evaluator_runs_for_eval_job(
                    metrics, evaluator_id_by_metric_key
                )
                runs = [r.model_dump() for r in built] if built else None
                if runs:
                    pr["evaluator_runs"] = runs
            except Exception as e:
                logger.warning(f"post_process: failed to build evaluator_runs: {e}")

        # Backfill evaluator_version_id / output_type onto each run entry from
        # the snapshot — the metric-key map gives us UUIDs but not types.
        for run in pr.get("evaluator_runs") or []:
            if not isinstance(run, dict):
                continue
            uid = run.get("evaluator_uuid")
            snap = snapshot_by_uuid.get(uid) if uid else None
            if snap:
                # `setdefault` won't overwrite an existing-but-None value
                # (the Pydantic model dumps these as None when unset), so
                # explicitly fill any None slot from the snapshot.
                if run.get("evaluator_version_id") is None:
                    run["evaluator_version_id"] = snap.get("evaluator_version_id")
                if run.get("output_type") is None:
                    run["output_type"] = snap.get("output_type")
                if run.get("name") is None:
                    run["name"] = snap.get("name")

        # ---------- Steps 3 + 4 + bonus: per-row evaluator_outputs ----------
        rows = pr.get("results") or []
        if not rows:
            continue

        # Build the {column_name -> snapshot_dict} pairs we'll lift each row
        # against. Prefer the calibrate-authoritative `evaluators_map` when
        # available — it's the only string we *know* identifies an evaluator
        # output column rather than a built-in row column. Fall back to the
        # snapshot's display name when the map is missing (in-progress reads
        # before config.json lands, or legacy pre-map data); that fallback
        # is vulnerable to reserved-column / duplicate-name collisions, so
        # it's accepted only when there's no better option.
        if evaluator_id_by_metric_key:
            lift_pairs = [
                (col_name, snapshot_by_uuid.get(uid))
                for col_name, uid in evaluator_id_by_metric_key.items()
            ]
        else:
            lift_pairs = list(snapshot_by_name.items())

        for row in rows:
            if not isinstance(row, dict):
                continue
            outputs = row.get("evaluator_outputs")
            if not isinstance(outputs, dict):
                outputs = {}

            for column_name, snap in lift_pairs:
                if not snap:
                    continue
                uid = snap.get("uuid") or snap.get("evaluator_id")
                if not uid:
                    continue
                # Calibrate column names: "<column>" carries the value,
                # "<column>_reasoning" carries the judge's free-text rationale.
                if column_name not in row and f"{column_name}_reasoning" not in row:
                    continue
                raw_value = row.get(column_name)
                reasoning = row.get(f"{column_name}_reasoning")
                is_err = _row_value_looks_like_error(raw_value, reasoning)

                if is_err:
                    typed_value: Any = None
                else:
                    output_type = (snap.get("output_type") or "").lower()
                    if output_type in ("binary", "rating"):
                        typed_value = coerce_evaluator_score(raw_value, output_type)
                    else:
                        # Unknown/unspecified — leave as-is so we don't lossy-
                        # convert future evaluator types we haven't met.
                        typed_value = raw_value

                entry: Dict[str, Any] = {
                    "value": typed_value,
                    "reasoning": reasoning,
                    "evaluator_version_id": snap.get("evaluator_version_id"),
                    "output_type": snap.get("output_type"),
                    # `name` is the human-facing display name from the
                    # snapshot, not the calibrate column key — they're
                    # equal in practice but we surface the snapshot's
                    # name so the FE shows what the user sees in the
                    # evaluator list, not calibrate's internal key.
                    "name": snap.get("name") or column_name,
                }
                if is_err:
                    entry["error"] = True
                outputs[uid] = entry

            if outputs:
                row["evaluator_outputs"] = outputs

            # Coerce known numeric row columns (wer/cer/...) from string→float
            # so the FE doesn't need parallel coercion paths.
            for key in list(row.keys()):
                if key in _NUMERIC_ROW_KEYS:
                    row[key] = _coerce_numeric(row[key])


def load_evaluator_metric_key_map(details: Optional[Dict[str, Any]]) -> Dict[str, str]:
    """Read the calibrate-side ``{metric_key: evaluator_uuid}`` map from a
    job's on-disk ``output_dir/config.json``. Returns ``{}`` if the dir is
    missing or the read fails — both are normal mid-flight states (job
    queued, output_dir already cleaned up, etc.) and should not throw.

    Extracted to share between STT and TTS GET handlers, which both need
    this map to drive ``post_process_provider_results`` mid-flight.
    """
    if not details:
        return {}
    output_dir_str = details.get("output_dir")
    if not output_dir_str:
        return {}
    try:
        candidate = Path(output_dir_str)
        if not candidate.exists():
            return {}
        return read_evaluators_map_from_config(candidate)
    except Exception:
        return {}


def compute_share_token_toggle(
    job: Optional[Dict[str, Any]],
    is_public: bool,
    *,
    token_field: str = "share_token",
    token_factory: Optional[Any] = None,
) -> tuple:
    """Shared logic for every visibility-toggle PATCH endpoint.

    Returns ``(token_to_persist, token_to_return)``:

    * ``token_to_persist`` keeps any existing token across off→on→off
      cycles so a previously-distributed share URL keeps working when
      sharing is re-enabled. (Codex P2 caught the bug where 3/4
      handlers were NULLing the token on disable.) Lookup queries
      already filter on ``is_public = 1``, so a stored-but-disabled
      token cannot resolve.
    * ``token_to_return`` is suppressed (None) when sharing is off so
      the FE never displays a share URL while the link is dead.

    ``token_factory`` defaults to ``str(uuid.uuid4())`` for parity with
    the historical STT/TTS/annotation-eval shapes. Pass
    ``secrets.token_urlsafe(24)`` for the labelling-job ``view_token``.
    """
    if token_factory is None:
        import uuid as _uuid

        token_factory = lambda: str(_uuid.uuid4())

    existing = (job or {}).get(token_field)
    if is_public:
        token_to_persist = existing or token_factory()
    else:
        token_to_persist = existing
    return token_to_persist, (token_to_persist if is_public else None)


def enrich_evaluator_runs_with_current_names(
    provider_results: Optional[List[Any]],
    evaluator_snapshots: Optional[List[Dict[str, Any]]] = None,
) -> None:
    """Mutates each provider result dict in place and refreshes evaluator run metadata."""
    if not provider_results:
        return
    from db import get_evaluator

    snapshot_by_uuid = {
        s["uuid"]: s
        for s in (evaluator_snapshots or [])
        if isinstance(s, dict) and s.get("uuid")
    }

    for pr in provider_results:
        runs_raw = pr.get("evaluator_runs")
        if not runs_raw:
            continue
        for run in runs_raw:
            if not isinstance(run, dict):
                continue
            uid = run.get("evaluator_uuid")
            row = get_evaluator(uid) if uid else None
            snapshot = snapshot_by_uuid.get(uid) if uid else None
            run["name"] = (
                (row.get("name") if row else None)
                or (snapshot.get("name") if snapshot else None)
                or run.get("metric_key")
                or ""
            )
            run["description"] = row.get("description") if row else None


def read_leaderboard_xlsx(leaderboard_dir: Path) -> Optional[List[dict]]:
    """Read the leaderboard summary from the xlsx file in leaderboard directory.

    Looks for any .xlsx file in the directory (commonly stt_leaderboard.xlsx or tts_leaderboard.xlsx).
    """
    if not leaderboard_dir.exists():
        logger.warning(f"Leaderboard directory does not exist: {leaderboard_dir}")
        return None

    xlsx_files = list(leaderboard_dir.glob("*.xlsx"))
    if not xlsx_files:
        logger.warning(
            f"No xlsx files found in leaderboard directory: {leaderboard_dir}"
        )
        all_files = list(leaderboard_dir.iterdir())
        logger.info(f"Files in leaderboard directory: {[f.name for f in all_files]}")
        return None

    xlsx_file = xlsx_files[0]
    logger.info(f"Reading leaderboard from: {xlsx_file}")

    try:
        wb = openpyxl.load_workbook(str(xlsx_file), data_only=True)
        logger.info(f"Workbook sheets: {wb.sheetnames}")

        if "summary" not in wb.sheetnames:
            logger.warning(
                f"'summary' sheet not found in {xlsx_file.name}, sheets: {wb.sheetnames}"
            )
            return None

        ws = wb["summary"]
        headers = [cell.value for cell in ws[1] if cell.value is not None]
        logger.info(f"Leaderboard headers: {headers}")

        leaderboard_summary = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            if any(cell.value is not None for cell in row):
                row_dict = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers):
                        row_dict[headers[idx]] = cell.value
                if any(v is not None for v in row_dict.values()):
                    leaderboard_summary.append(row_dict)

        logger.info(f"Read {len(leaderboard_summary)} rows from leaderboard")
        return leaderboard_summary
    except Exception as e:
        logger.warning(f"Failed to read leaderboard xlsx: {e}")
        return None
